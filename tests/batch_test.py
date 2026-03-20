"""
tests/batch_test.py — Batch test + automatický export do eval_template.xlsx

Spustenie (z koreňa projektu):
    python tests/batch_test.py

Čo robí:
  1. Pre každú otázku zavolá ask() — retrieval + LLM
  2. Výsledky automaticky zapíše do tests/results/eval_template.xlsx
     (stĺpce D=odpoveď, E=similarity, F=čas, J=§citovaný, K=zákon, L=KW ratio, M=KW chýba, N=auto_score, S=zdroje)
  3. Stĺpce O/P/Q (manual_correct, hallucination, komentár) ostávajú prázdne — vyplní Opus alebo Pavel

Prerekvizita: Ollama musí bežať na localhost:11434

Ak chýba openpyxl:
    pip install openpyxl
"""

import sys
import json
import time
import subprocess
from pathlib import Path
from datetime import datetime

# Auto-install openpyxl ak chýba
try:
    import openpyxl  # noqa: F401
except ImportError:
    print("Inštalujem openpyxl...")
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
    print("openpyxl nainštalovaný.")

# Pridaj rodičovský adresár do path
sys.path.insert(0, str(Path(__file__).parent.parent))

from rag import ask

# ── Cesty ─────────────────────────────────────────────────────────────────────
BASE        = Path(__file__).parent
QUESTIONS_F = BASE / "test_questions.json"
EXCEL_F     = BASE / "results" / "eval_template.xlsx"
(BASE / "results").mkdir(exist_ok=True)

# ── Helpers ───────────────────────────────────────────────────────────────────
def check_keywords(answer: str, keywords: list) -> dict:
    al = answer.lower()
    found   = [k for k in keywords if k.lower() in al]
    missing = [k for k in keywords if k.lower() not in al]
    return {"found": found, "missing": missing,
            "ratio": round(len(found)/len(keywords), 2) if keywords else 1.0}

def check_par(answer: str, exp: str) -> str:
    if not exp:
        return "N/A"
    return "ÁNO" if exp.lower() in answer.lower() else "NIE"

def check_zakon(answer: str, exp: str) -> str:
    if not exp:
        return "N/A"
    cislo = exp.split("/")[0]
    return "ÁNO" if cislo in answer else "NIE"

# ── Hlavný beh ────────────────────────────────────────────────────────────────
def run():
    with open(QUESTIONS_F, encoding="utf-8") as f:
        questions = json.load(f)

    print(f"Batch test: {len(questions)} otázok | {datetime.now().strftime('%H:%M:%S')}")
    print(f"Excel výstup: {EXCEL_F}")
    print("=" * 70)

    rows = []
    for i, q in enumerate(questions, 1):
        print(f"[{i:02d}/{len(questions)}] {q['id']}  {q['question'][:55]}...", end=" ", flush=True)
        t0 = time.time()
        try:
            result = ask(q["question"])
            elapsed = round(time.time() - t0, 1)

            kw = check_keywords(result["answer"], q.get("expected_keywords", []))
            par_ok  = check_par(result["answer"],   q.get("expected_paragraf", ""))
            zak_ok  = check_zakon(result["answer"],  q.get("expected_zakon", ""))

            auto = kw["ratio"]
            if par_ok  == "NIE": auto = max(0.0, auto - 0.3)
            if zak_ok  == "NIE": auto = max(0.0, auto - 0.2)
            auto = round(auto, 2)

            rows.append({
                "id":       q["id"],
                "answer":   result["answer"],
                "sim":      result["similarity_max"],
                "time_s":   elapsed,
                "par_ok":   par_ok,
                "zak_ok":   zak_ok,
                "kw_ratio": kw["ratio"],
                "kw_miss":  ", ".join(kw["missing"]),
                "auto":     auto,
                "sources":  " | ".join(result["sources"]),
                "below":    result.get("below_threshold", False),
            })
            status = "✓" if auto >= 0.6 else "✗"
            print(f"{status}  sim={result['similarity_max']:.0%}  auto={auto:.2f}  {elapsed}s")
        except Exception as e:
            elapsed = round(time.time() - t0, 1)
            rows.append({"id": q["id"], "answer": f"ERROR: {e}",
                         "sim": 0, "time_s": elapsed, "par_ok": "ERR",
                         "zak_ok": "ERR", "kw_ratio": 0, "kw_miss": "", "auto": 0, "sources": ""})
            print(f"✗  CHYBA: {e}")

    # ── Najprv uloz JSON (nezavisi od openpyxl) ───────────────────────────────
    import json as _json
    json_out = BASE / "results" / f"batch_raw_{datetime.now().strftime('%Y%m%d_%H%M')}.json"
    with open(json_out, "w", encoding="utf-8") as jf:
        _json.dump(rows, jf, ensure_ascii=False, indent=2)
    print(f"\n✓ JSON záloha: {json_out}")

    # ── Zapis do Excelu ───────────────────────────────────────────────────────
    _write_excel(rows, questions)

    # ── Suhrn ─────────────────────────────────────────────────────────────────
    scores = [r["auto"] for r in rows]
    print("\n" + "=" * 70)
    print(f"Priemerný auto_score: {sum(scores)/len(scores):.2f}")
    print(f"Auto ✓ (>=0.6): {sum(1 for s in scores if s>=0.6)}/{len(scores)}")
    print(f"Auto ✗ (<0.6):  {sum(1 for s in scores if s<0.6)}/{len(scores)}")
    print(f"\nExcel: {EXCEL_F}")
    print("Ďalší krok: otvor Excel, skontroluj stĺpec D (odpovede LLM)")


def _write_excel(rows: list, questions: list):
    """Zapíše výsledky do existujúceho eval_template.xlsx."""
    from openpyxl import load_workbook

    if not EXCEL_F.exists():
        print(f"WARN: {EXCEL_F} neexistuje — vygeneruj ho najprv cez make_eval_xlsx.py")
        _write_csv_fallback(rows)
        return

    wb = load_workbook(EXCEL_F)
    ws = wb["Eval"]

    # Mapovanie stĺpcov (písmeno → kľúč v rows dict)
    COL_MAP = {
        "D": "answer",
        "E": "sim",
        "F": "time_s",
        "J": "par_ok",
        "K": "zak_ok",
        "L": "kw_ratio",
        "M": "kw_miss",
        "N": "auto",
        "S": "sources",
    }

    from openpyxl.styles import Font, Alignment, PatternFill
    NORM = Font(name="Arial", size=9)
    WRAP = Alignment(wrap_text=True, vertical="top")
    RED_FILL  = PatternFill("solid", start_color="FFD7D7")
    GREEN_FILL= PatternFill("solid", start_color="D7FFD7")

    for i, (row_data, q) in enumerate(zip(rows, questions)):
        excel_row = i + 4  # data začína na riadku 4

        for col, key in COL_MAP.items():
            val = row_data.get(key, "")
            cell = ws[f"{col}{excel_row}"]
            cell.value = val
            cell.font  = NORM
            cell.alignment = WRAP

            # Farebné zvýraznenie pre par_ok / zak_ok
            if key in ("par_ok", "zak_ok"):
                if val == "NIE":
                    cell.fill = RED_FILL
                elif val == "ÁNO":
                    cell.fill = GREEN_FILL

            # Farebné zvýraznenie pre auto score
            if key == "auto":
                if isinstance(val, float) and val < 0.6:
                    cell.fill = RED_FILL
                elif isinstance(val, float) and val >= 0.8:
                    cell.fill = GREEN_FILL

            # Formát percent pre sim
            if key == "sim":
                cell.number_format = "0%"

    wb.save(EXCEL_F)
    print(f"\n✓ Excel aktualizovaný: {EXCEL_F}")


def _write_csv_fallback(rows: list):
    """Záložný CSV export ak Excel neexistuje."""
    import csv
    csv_out = BASE / "results" / f"batch_fallback_{datetime.now().strftime('%Y%m%d_%H%M')}.csv"
    with open(csv_out, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=rows[0].keys())
        writer.writeheader()
        writer.writerows(rows)
    print(f"CSV záloha: {csv_out}")


if __name__ == "__main__":
    run()
